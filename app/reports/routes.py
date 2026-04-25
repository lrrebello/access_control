from flask import render_template, request, send_file, flash, redirect, url_for, current_app, render_template_string
from app import db
from app.models import AccessLog
from app.reports import reports
from flask_login import login_required, current_user
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO
import os
from PIL import Image

# Mapeamento seguro para ordenação
SORT_MAPPING = {
    'vehicle_plate': AccessLog.vehicle_plate,
    'trailer_plate': AccessLog.trailer_plate,
    'driver_name': AccessLog.driver_name,
    'company': AccessLog.company,
    'entry_time': AccessLog.entry_time,
    'exit_time': AccessLog.exit_time,
    'vehicle_type': AccessLog.vehicle_type,
}

@reports.route("/reports", methods=['GET', 'POST'])
@login_required
def view_reports():
    # Ordenação
    sort_by = request.args.get('sort_by', 'entry_time')
    sort_dir = request.args.get('sort_dir', 'desc')

    if sort_by not in SORT_MAPPING:
        sort_by = 'entry_time'
    if sort_dir not in ['asc', 'desc']:
        sort_dir = 'desc'

    # Filtros de data
    start_date_str = request.form.get('start_date') or request.args.get('start_date')
    end_date_str = request.form.get('end_date') or request.args.get('end_date')

    query = AccessLog.query
    
    # FILTRO POR USUÁRIO - ADMINS veem tudo, outros veem só seus registros
    if not current_user.is_admin:
        query = query.filter(AccessLog.user_id == current_user.id)

    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        query = query.filter(AccessLog.entry_time >= start_date)
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(AccessLog.entry_time < end_date)

    # Aplicar ordenação
    order_column = SORT_MAPPING[sort_by]
    query = query.order_by(order_column.desc() if sort_dir == 'desc' else order_column.asc())

    logs = query.all()

    # Exportações
    if request.method == 'POST':
        if 'export_excel' in request.form:
            return export_excel(logs)
        if 'export_pdf' in request.form:
            return export_pdf(logs, start_date=start_date_str, end_date=end_date_str)

    return render_template('reports/reports.html',
                           logs=logs,
                           title='Relatórios',
                           start_date=start_date_str,
                           end_date=end_date_str,
                           sort_by=sort_by,
                           sort_dir=sort_dir)


def export_excel(logs):
    """Exporta para Excel com coluna de acompanhantes e posto"""
    data = []
    for log in logs:
        # Formatar lista de acompanhantes
        companions_text = ""
        if log.companions:
            companions_list = [f"{c.name} ({c.document})" for c in log.companions]
            companions_text = "\n".join(companions_list)
        
        # Nome do posto (se existir)
        posto = log.workstation.name if log.workstation else 'Não informado'
        
        data.append({
            'Posto': posto,
            'Matrícula': log.vehicle_plate,
            'Reboque': log.trailer_plate or '',
            'Tipo Veículo': log.vehicle_type,
            'Condutor': log.driver_name,
            'Doc. Condutor': log.driver_doc,
            'Acompanhantes': companions_text,
            'Qtd. Acomp.': len(log.companions),
            'Total Pessoas': log.total_people,
            'Empresa': log.company,
            'Entrada': log.entry_time.strftime('%d/%m/%Y %H:%M'),
            'Saída': log.exit_time.strftime('%d/%m/%Y %H:%M') if log.exit_time else 'Em local',
            'Tempo Permanência': log.duration if log.exit_time else 'Em local',
            'Observação': log.observations or '',
            'Alerta': log.alert_msg or ''
        })
    
    df = pd.DataFrame(data)
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Relatório de Acessos')
        
        # Acessar a planilha
        worksheet = writer.sheets['Relatório de Acessos']
        
        # Adicionar cabeçalho com informações do usuário e posto
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        
        # Determinar o posto atual do usuário (se houver)
        posto_atual = current_user.current_workstation.name if current_user.current_workstation else 'Todos os Postos'
        
        # Inserir linhas no topo
        worksheet.insert_rows(0, 5)
        
        # Título do relatório
        worksheet['A1'] = 'RELATÓRIO DE CONTROLE DE ACESSO'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet.merge_cells('A1:O1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Informações do relatório
        worksheet['A2'] = f'Posto: {posto_atual}'
        worksheet['A3'] = f'Gerado por: {current_user.username}'
        worksheet['A4'] = f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        worksheet['A5'] = f'Período: {request.form.get("start_date", "Tudo")} até {request.form.get("end_date", "Tudo")}'
        
        # Ajustar largura das colunas
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if hasattr(cell, 'value') and cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    output.seek(0)
    filename = f"relatorio_acessos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def export_pdf(logs, start_date=None, end_date=None, is_today_exits=False):
    """Gera PDF com acompanhantes e posto"""
    from xhtml2pdf import pisa
    from io import BytesIO
    from PIL import Image
    import sys
    
    # Caminho da logo
    logo_path = os.path.join(current_app.root_path, 'static', 'logo.png')
    resized_logo_path = os.path.join(current_app.root_path, 'static', 'logo_resized.png')

    try:
        if os.path.exists(logo_path):
            img = Image.open(logo_path).convert("RGBA")
            img.thumbnail((220, 60), Image.Resampling.LANCZOS)
            img.save(resized_logo_path, format='PNG', optimize=True)
            logo_to_use = resized_logo_path
        else:
            logo_to_use = ""
    except Exception:
        logo_to_use = ""

    try:
        # Escolher o template baseado no tipo de relatório
        if is_today_exits:
            template_file = 'reports/pdf_today_exits_template.html'
            titulo = "RELATÓRIO DE SAÍDAS DO DIA"
        else:
            template_file = 'reports/pdf_template.html'
            titulo = "RELATÓRIO DE CONTROLE DE ACESSO"
        
        html_content = open(f'app/templates/{template_file}', encoding='utf-8').read()
        
        # Dados do usuário
        user_data = {
            'username': str(current_user.username),
            'is_admin': bool(current_user.is_admin)
        }
        
        # Posto atual do usuário
        user_posto = current_user.current_workstation.name if current_user.current_workstation else 'Todos os Postos'
        
        # Processar logs
        logs_data = []
        for log in logs:
            log_dict = {
                'vehicle_plate': str(log.vehicle_plate),
                'trailer_plate': str(log.trailer_plate) if log.trailer_plate else '',
                'vehicle_type': str(log.vehicle_type),
                'driver_name': str(log.driver_name),
                'driver_doc': str(log.driver_doc),
                'company': str(log.company),
                'entry_time': log.entry_time,
                'exit_time': log.exit_time,
                'observations': str(log.observations) if log.observations else '',
                'duration': str(log.duration) if log.duration else '',
                'alert_msg': str(log.alert_msg) if log.alert_msg else '',
                'workstation': {'name': log.workstation.name if log.workstation else '-'},
                'companions': []
            }
            
            for companion in log.companions:
                log_dict['companions'].append({
                    'name': str(companion.name),
                    'document': str(companion.document)
                })
            
            logs_data.append(log_dict)
        
        html = render_template_string(
            html_content,
            logs=logs_data,
            now=datetime.now(),
            start_date=start_date or '',
            end_date=end_date or '',
            logo_path=logo_to_use,
            user=user_data,
            user_posto=user_posto,
            titulo=titulo
        )

        output = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), dest=output)

        if pdf.err:
            flash('Erro ao gerar o PDF.', 'danger')
            return redirect(url_for('reports.view_reports'))

        output.seek(0)
        
        if is_today_exits:
            filename = f"saidas_hoje_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        else:
            filename = f"relatorio_acessos_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/pdf')

    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'danger')
        return redirect(url_for('reports.view_reports'))
    
@reports.route("/export/today-exits-pdf")
@login_required
def export_today_exits_pdf():
    """Exporta PDF apenas com as saídas registradas hoje"""
    today = datetime.now().date()
    today_start = datetime(today.year, today.month, today.day, 0, 0, 0)
    today_end = datetime(today.year, today.month, today.day, 23, 59, 59)
    
    query = AccessLog.query
    
    # Filtrar por posto de trabalho
    if current_user.active_workstation_id:
        query = query.filter(AccessLog.workstation_id == current_user.active_workstation_id)
    
    # Apenas saídas de hoje
    query = query.filter(
        AccessLog.exit_time != None,
        AccessLog.exit_time >= today_start,
        AccessLog.exit_time <= today_end
    )
    
    # Admins veem tudo, outros veem só seus registros
    if not current_user.is_admin:
        query = query.filter(AccessLog.user_id == current_user.id)
    
    logs = query.order_by(AccessLog.exit_time.desc()).all()
    
    if not logs:
        flash('Nenhuma saída registrada hoje para gerar o PDF.', 'warning')
        # Voltar para o dashboard com o filtro das saídas de hoje
        return redirect(url_for('main.dashboard', filter='today_exits'))
    
    return export_pdf(logs, 
                      start_date=today.strftime('%d/%m/%Y'), 
                      end_date=today.strftime('%d/%m/%Y'), 
                      is_today_exits=True)